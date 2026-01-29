import copy
import os.path
import re
from django.conf import settings
from django.http import FileResponse, HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from .models import Registry, Act, Conformity, Boss, StickPlace, Manufacturer, RegistryModify
import openpyxl
from docxtpl import DocxTemplate
from django.contrib.auth.decorators import permission_required
from distribution.models import ControlJournal
from application.models import Status, Application
from .forms import ActInput, ActDataInput
from django.forms import formset_factory
from datetime import date

color_mapp = {
        'ЕАН': 'table-success',
        'СГА': 'table-info',
        'МАА': 'table-warning',
        'АВВ': 'table-danger'
    }


@permission_required('act_creation.view_registry')
def registry(request):
    registries = Registry.objects.all()
    list_keys = ['number', 'model', 'version', 'manufacturer']
    param_dict = {key : request.GET.get(key) for key in list_keys}
    for param, value in param_dict.items():
        if value:
            filter_name = f'{param}__icontains'
            registries = registries.filter(**{filter_name: value})
    context = {
        'registries': registries,
        'number_get': param_dict['number'],
        'model_get': param_dict['model'],
        'version_get': param_dict['version'],
        'manufacturer_get': param_dict['manufacturer']
         }
    return render(request, 'registry.html', context)


def registry_input(request):
    wb = openpyxl.load_workbook(
        'D:\PythonCreateActTOProject\Reestr\Реестр ИА.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        registry_new = Registry(
            number=row[0].value,
            model=row[1].value,
            version=row[2].value,
            manufacturer=row[3].value
        )
        registry_new.save()
    registries = Registry.objects.all()
    context = {
            'registries': registries
        }
    return render(request, 'registry.html', context)

@permission_required('act_creation.view_registry')
def registry_modify(request):
    registries = RegistryModify.objects.all()
    list_keys = ['number', 'model', 'version', 'manufacturer']
    param_dict = {key : request.GET.get(key) for key in list_keys}
    for param, value in param_dict.items():
        if value:
            filter_name = f'{param}__icontains'
            registries = registries.filter(**{filter_name: value})
    context = {
        'registries': registries,
        'number_get': param_dict['number'],
        'model_get': param_dict['model'],
        'version_get': param_dict['version'],
        'manufacturer_get': param_dict['manufacturer']
         }
    return render(request, 'registry_modify.html', context)


def registry_modify_input(request):
    wb = openpyxl.load_workbook(
        'D:\PythonCreateActTOProject\Reestr\Реестр ИА по версиям 2025_12_02.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        registry_new = RegistryModify(
            number=row[0].value,
            model=row[1].value,
            version=row[2].value,
            manufacturer=row[3].value
        )
        registry_new.save()
    registries = RegistryModify.objects.all()
    context = {
            'registries': registries
        }
    return render(request, 'registry_modify.html', context)

@permission_required('act_creation.view_act')
def slot_machine_data(request):
    slot_machines_data = Act.objects.all().order_by('-act_number')
    list_keys = [
        'act_number', 'application', 'control_sticks_number',
        'declarant', 'result', 'model', 'version', 'slot_number',
        'reg_number', 'board_number'
    ]
    param_dict = {_: request.GET.get(_) for _ in list_keys}
    for key, value in param_dict.items():
        if value:
            if key == 'application':
                filter_name = f'distribution__application__application_number__icontains'
            elif key == 'declarant':
                filter_name = f'distribution__application__declarant__name__icontains'
            elif key == 'model' or key == 'version':
                filter_name = f'model_registry__{key}__icontains'
            elif key == 'reg_number':
                filter_name = f'model_registry__number__icontains'
            else:
                filter_name = f'{key}__icontains'
            slot_machines_data = slot_machines_data.filter(**{filter_name: str(value)})

    context = {
            'slot_machines_data': slot_machines_data,
            'act_number': param_dict['act_number'],
            'application': param_dict['application'],
            'control_sticks_number': param_dict['control_sticks_number'],
            'declarant': param_dict['declarant'],
            'result': param_dict['result'],
            'model': param_dict['model'],
            'version': param_dict['version'],
            'slot_number': param_dict['slot_number'],
            'reg_number': param_dict['reg_number'],
            'board_number': param_dict['board_number'],
            'color_mapp': color_mapp
        }

    return render(request, 'slot_machine_data.html', context)


def slot_machine_data_change(request, pk):
    slot_machine_data_change = Act.objects.get(pk=pk)
    success_message = ''
    if request.method == 'POST':
        form = ActInput(request.POST, instance=slot_machine_data_change)
        if form.is_valid():
            form.save()
            success_message = 'Данные по ИА успешно изменены!'
    else:
        form = ActInput(instance=slot_machine_data_change)
    context = (
        {
            'form': form,
            'slot_machine_data': slot_machine_data_change,
            'success_message': success_message
        }
    )
    return render(request, 'slot_machine_data_change.html', context)


@permission_required('act_creation.add_act')
def s_m_data_input(request, pk):
    """Представление для занесения данных об игровом автомате в журнал актов ИА"""
    users_application = ControlJournal.objects.get(pk=pk)
    slot_machines_data = Act.objects.all().order_by('-act_number')
    conformity = Conformity.objects.all()
    registry = RegistryModify.objects.all()
    ActFormSet = formset_factory(ActInput,extra=1)
    success_message = ''
    number_stickers = '1'

    if request.method == 'GET':
        data_old = request.session.get('act_formset_data')
        data = copy.deepcopy(data_old)
        number_stickers = request.session.get('number_stickers', '1')
        if data:
            initial_session_data = data
            for form in initial_session_data:
                form['conformity'] = conformity.get(id=form['conformity'])
                form['model_registry'] = registry.get(id=form['model_registry'])
        else:
            initial_session_data = None
        formset = ActFormSet(initial=initial_session_data)

        list_keys = [
            'act_number', 'application', 'control_sticks_number',
            'declarant', 'result', 'model', 'version', 'slot_number',
            'reg_number', 'board_number'
        ]
        param_dict = {_: request.GET.get(_, '') for _ in list_keys}
        request.session['filters'] = param_dict

    elif request.method == 'POST':
        param_dict = request.session.get('filters', '')
        formset = ActFormSet(request.POST)
        if 'get_data_to_formset' in request.POST.keys():
            pk_data = request.POST.get('get_data_to_formset')
            data_to_formset = Act.objects.get(pk=pk_data)
            data = request.session.get('act_formset_data')
            new_data_dict = {
                'act_number': None,
                'control_sticks_number': '',
                'conformity': data_to_formset.conformity.id,
                'model_registry': data_to_formset.model_registry.id,
                'slot_number': data_to_formset.slot_number,
                'board_number': data_to_formset.board_number
            }
            if data:
                data.append(new_data_dict)
            else:
                data = list()
                data.append(new_data_dict)
            request.session['act_formset_data'] = data
            number_stickers = request.session.get('number_stickers', '1')
            formset = ActFormSet(initial=data)
            success_message = f'Данные ИА с номером {data_to_formset.slot_number} добавлены в форму и сохранены в сессии.'
        else:
            number_stickers = request.POST['amount_numbers']
            request.session['number_stickers'] = number_stickers
            if formset.is_valid():
                if 'save_draft' in request.POST.keys(): # сохранение черновика
                    cleaned_data_list = [form.cleaned_data for form in formset if form.cleaned_data]
                    cleaned_data_list = [form for form in cleaned_data_list if not form['is_deleted']]
                    for form in cleaned_data_list:
                        pattern = r'^\d{6}$'
                        form_stickers = form['control_sticks_number']
                        if re.match(pattern, form_stickers):
                            num = int(form_stickers)
                            nums = [str(i + num) for i in range(int(number_stickers))]
                            form['control_sticks_number'] = ' '.join(nums)
                        form['conformity'] = form['conformity'].id
                        form['model_registry'] = form['model_registry'].id
                    request.session['act_formset_data'] = cleaned_data_list
                    formset = ActFormSet(initial=cleaned_data_list)
                    success_message = 'Черновик успешно сохранен в сессии.'
                elif 'save_form' in request.POST.keys(): # отправка данных об ИА в журнал актов ('Act')
                    last_act = slot_machines_data.first().act_number
                    formset = ActFormSet(request.POST)
                    if formset.is_valid():
                        act_number = None
                        for form in formset:
                            if form.cleaned_data and form.cleaned_data.get('act_number'):
                                act_number = form.cleaned_data.get('act_number')
                        if act_number is None:
                            act_number = last_act + 1
                        for form in formset:
                            if form.cleaned_data:
                                num_SK = form.cleaned_data.get('control_sticks_number').split()
                                control_sticks_number_corrected = [f'ИА {num}' for num in num_SK if num.isdigit()]
                                control_sticks_number_cleaned = ' '.join(control_sticks_number_corrected)
                                user_now = request.user
                                slot_machines_data_new = Act(
                                    act_number=act_number,
                                    distribution=users_application,
                                    control_sticks_number=control_sticks_number_cleaned,
                                    conformity=form.cleaned_data['conformity'],
                                    model_registry=form.cleaned_data['model_registry'],
                                    slot_number=form.cleaned_data['slot_number'],
                                    board_number=form.cleaned_data['board_number'],
                                    user=user_now
                                )
                                slot_machines_data_new.save()
                                formset = ActFormSet()
                                request.session.pop('act_formset_data', None)
                                request.session.pop('number_stickers', None)
                                success_message = 'Акт успешно добавлен в журнал'

    for key, value in param_dict.items():
        if value:
            if key == 'application':
                filter_name = f'distribution__application__application_number__icontains'
            elif key == 'declarant':
                filter_name = f'distribution__application__declarant__name__icontains'
            elif key == 'result':
                filter_name = f'conformity__conformity__icontains'
            elif key == 'model' or key == 'version':
                filter_name = f'model_registry__{key}__icontains'
            elif key == 'reg_number':
                filter_name = f'model_registry__number__icontains'
            else:
                filter_name = f'{key}__icontains'
            slot_machines_data = slot_machines_data.filter(**{filter_name: str(value)})
    slot_machines_count = slot_machines_data.filter(
        distribution__application__application_number=str(users_application)).count()
    context = (
        {
            'slot_machines_data': slot_machines_data,
            'slot_machines_count': slot_machines_count,
            'users_application': users_application,
            'formset': formset,
            'success_message': success_message,
            'act_number': param_dict['act_number'],
            'application': param_dict['application'],
            'control_sticks_number': param_dict['control_sticks_number'],
            'declarant': param_dict['declarant'],
            'result': param_dict['result'],
            'model': param_dict['model'],
            'version': param_dict['version'],
            'slot_number': param_dict['slot_number'],
            'reg_number': param_dict['reg_number'],
            'board_number': param_dict['board_number'],
            'empty_form': formset.empty_form,
            'number_stickers': number_stickers,
            'color_mapp': color_mapp
        }
    )
    return render(request, 's_m_data_input.html', context)

@permission_required('act_creation.view_act')
def act_creation(request):
    """Страница с заявками конкретного исполнителя"""
    user_now = request.user
    users_applications = ControlJournal.objects.filter(user_many__in=[user_now], application__status='4').order_by('-id')
    context = {
            'users_applications': users_applications,
            'user_now': user_now
        }
    return render(request, 'act_creation.html', context)

@permission_required('act_creation.add_act')
def docx_create(request):
    """Страница для создания акта ТО"""
    slot_machines_data = Act.objects.all().order_by('-act_number')
    bosses = Boss.objects.all()
    stick_places = StickPlace.objects.all()
    manufacturers = Manufacturer.objects.all()
    success_message, danger_message = '', ''
    monthes = {1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
               7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'}
    date_now = date.today()
    date_filename = date_now.strftime("%d.%m.%y")
    year = date_now.strftime("%Y")
    date_for_word = f'{date_now.strftime("%d")} {monthes[date_now.month]} {year} г.'
    form = ActDataInput()
    if request.method == 'POST':
        form = ActDataInput(request.POST)
        if form.is_valid():
            acts_query = slot_machines_data.filter(act_number=int(str(form.cleaned_data['act_number'])))
            act_first = acts_query.first()
            bill_date = act_first.distribution.application.bill_date.strftime("%d.%m.%Y")
            bill_num = f'{act_first.distribution.application.bill_number} от {bill_date}'
            app_date = act_first.distribution.application.created_on.strftime("%d.%m.%Y")
            app_num = f'{act_first.distribution.application.application_number} от {app_date}'
            boss = bosses.get(name=str(form.cleaned_data['boss']))
            executor = f'{act_first.user.last_name} {act_first.user.first_name}'
            stick_place = stick_places.get(board_name=str(form.cleaned_data['stick_place']))
            manufacturer_form = manufacturers.get(name=str(form.cleaned_data['manufacturer']))
            if str(manufacturer_form) == 'Из реестра':
                manufacturer = act_first.model_registry.manufacturer
            else:
                manufacturer = manufacturer_form

            acts = []
            models_new = [act.model_registry.model for act in acts_query]
            versions_new = [act.model_registry.version for act in acts_query]
            model_reg_new = [act.model_registry.number.split()[0] for act in acts_query]
            slot_nums = [act.slot_number for act in acts_query]
            board_nums = [act.board_number for act in acts_query]
            acts_stickers = [act.control_sticks_number for act in acts_query]
            keys = ['model', 'version', 'model_registry', 'slot_number', 'board_number', 'sticks_number']
            lists = [models_new, versions_new, model_reg_new, slot_nums, board_nums, acts_stickers]
            for values in zip(*lists):
                act = dict(zip(keys, values))
                acts.append(act)

            word_context = {
                'act_number': act_first.act_number,
                'boss': boss.name,
                'boss_position': boss.position,
                'year': year,
                'date_for_word': date_for_word,
                'bill_num': bill_num,
                'app_num': app_num,
                'declarant': act_first.distribution.application.declarant.name,
                'declarant_address': act_first.distribution.application.declarant.address,
                'executor': executor,
                'sticker': stick_place.stick_place,
                'manufacturer': manufacturer,
                'acts': acts
            }
            docx_file_name = (f'{str(word_context['act_number'])} {word_context['declarant'].replace('"', '')} '
                      f'{stick_place.board_name} {date_filename}.docx')
            save_path = os.path.join(settings.MEDIA_ROOT, 'acts', docx_file_name)
            acts_len = len(acts)
            conformity = acts_query.first().conformity.id
            if conformity == 2:
                template_name = 'temp_negative.docx'
            elif stick_place == stick_places.get(pk=3) and 12 < acts_len <= 24: # разбиение контекста на страницы
                template_name = 'temp_tab_2.docx'
                word_context['acts_1'] = acts[:12]
                word_context['acts_2'] = acts[12:]
            elif stick_place == stick_places.get(pk=3) and 24 < acts_len <= 36: # CF1 pk=3
                template_name = 'temp_tab_3.docx'
                word_context['acts_1'] = acts[:12]
                word_context['acts_2'] = acts[12:24]
                word_context['acts_3'] = acts[24:]
            elif stick_place == stick_places.get(pk=3) and 36 < acts_len:
                template_name = 'temp_tab_4.docx'
                word_context['acts_1'] = acts[:12]
                word_context['acts_2'] = acts[12:24]
                word_context['acts_3'] = acts[24:36]
                word_context['acts_4'] = acts[36:]
            elif stick_place == stick_places.get(pk=4) and 24 < acts_len: # CF2 pk=4
                template_name = 'temp_tab_2.docx'
                word_context['acts_1'] = acts[:24]
                word_context['acts_2'] = acts[24:]
            elif stick_place == stick_places.get(pk=6) and 10 < acts_len: # EGT E3 pk=6
                template_name = 'temp_tab_2.docx'
                word_context['acts_1'] = acts[:10]
                word_context['acts_2'] = acts[10:]
            elif stick_place == stick_places.get(pk=7) and 8 < acts_len <= 16: # EGT E4 pk=7
                template_name = 'temp_tab_2.docx'
                word_context['acts_1'] = acts[:8]
                word_context['acts_2'] = acts[8:]
            elif stick_place == stick_places.get(pk=7) and 16 < acts_len: # EGT E4 pk=7
                template_name = 'temp_tab_3.docx'
                word_context['acts_1'] = acts[:8]
                word_context['acts_2'] = acts[8:16]
                word_context['acts_3'] = acts[16:]
            elif (stick_place == stick_places.get(pk=14) or stick_place == stick_places.get(pk=17)) and 6 < acts_len <= 12: # Promatic H2 и T3 pk=14,17
                template_name = 'temp_tab_2.docx'
                word_context['acts_1'] = acts[:6]
                word_context['acts_2'] = acts[6:]
            elif stick_place == stick_places.get(pk=10): #SET 5.7+ pk=10
                template_name = 'temp_SET.docx'
                stickers = [act.get('sticks_number') for act in acts][0]
                stickers_list = ['ИА ' + num for num in stickers.split() if num.isnumeric()]
                word_context['sticker_cupola'] = stickers_list[-1]
                word_context['sticker_terminal'] = stickers_list[-2]
                stickers_list_new = [' '.join(stickers_list[i:i+2]) for i in range(0,len(stickers_list) - 2,2)]
                word_context['stickers'] = stickers_list_new
            elif stick_place == stick_places.get(pk=15):  # SET 6.4+ pk=15
                template_name = 'temp_SET.docx'
                stickers = [act.get('sticks_number') for act in acts][0]
                stickers_list = ['ИА ' + num for num in stickers.split() if num.isnumeric()]
                word_context['sticker_cupola'] = stickers_list[-1]
                word_context['sticker_terminal'] = stickers_list[-2]
                stickers_list_new = [' '.join(stickers_list[i:i + 3]) for i in range(0, len(stickers_list) - 3, 3)]
                word_context['stickers'] = stickers_list_new
            elif stick_place == stick_places.get(pk=9): #ODREX pk=9
                template_name = 'temp_ODREX.docx'
                stickers = [act.get('sticks_number') for act in acts][0]
                stickers_list = ['ИА ' + num for num in stickers.split() if num.isnumeric()]
                word_context['sticker_cupola'] = stickers_list[-1]
                stickers_list_new = [' '.join(stickers_list[i:i+2]) for i in range(0,len(stickers_list) - 2,2)]
                word_context['stickers'] = stickers_list_new
            elif stick_place == stick_places.get(pk=13): #EGT BELL LINK pk=13
                template_name = 'temp_BELL_tab_1.docx'
                acts_stickers = [act.get('sticks_number') for act in acts]
                acts_stickers_mod = []
                jackpot_stickers = ''
                if acts_stickers:
                    for stickers in acts_stickers:
                        stickers_list = ['ИА ' + num for num in stickers.split() if num.isnumeric()]
                        stickers_list[-1] += '*'
                        stickers_list[-2] += '*'
                        jackpot_stickers = ', '.join(stickers_list[-2:])
                        stickers = ' '.join(stickers_list)
                        acts_stickers_mod.append(stickers)
                for key, value in enumerate(acts_stickers_mod):
                    acts[key]['sticks_number'] = value
                word_context['jackpot_stickers'] = jackpot_stickers
            else:
                template_name = 'temp_tab_1.docx'

            template_path = os.path.join(settings.BASE_DIR, 'templates', template_name)
            doc = DocxTemplate(template_path)
            doc.render(word_context)
            try:
                doc.save(save_path)
            except PermissionError:
                print('!!!ОШИБКА!!!')
                print(
                    'У вас открыт файл с таким же именем. Программа не может сохранить файл. Закройте файл word и повторите скрипт')
            success_message = f'Акт {docx_file_name} успешно сгенерирован и сохранен.'
            return redirect('download_act_docx', file_name=docx_file_name)

    list_keys = [
        'act_number', 'application', 'control_sticks_number',
        'declarant', 'result', 'model', 'version', 'slot_number',
        'reg_number', 'board_number'
    ]
    param_dict = {_: request.GET.get(_) for _ in list_keys}
    for key, value in param_dict.items():
        if value:
            if key == 'application':
                filter_name = f'distribution__application__application_number__icontains'
            elif key == 'declarant':
                filter_name = f'distribution__application__declarant__name__icontains'
            elif key == 'model' or key == 'version':
                filter_name = f'model_registry__{key}__icontains'
            elif key == 'reg_number':
                filter_name = f'model_registry__number__icontains'
            else:
                filter_name = f'{key}__icontains'
            slot_machines_data = slot_machines_data.filter(**{filter_name: str(value)})

    context = {
        'form': form,
        'slot_machines_data': slot_machines_data,
        'act_number': param_dict['act_number'],
        'application': param_dict['application'],
        'control_sticks_number': param_dict['control_sticks_number'],
        'declarant': param_dict['declarant'],
        'result': param_dict['result'],
        'model': param_dict['model'],
        'version': param_dict['version'],
        'slot_number': param_dict['slot_number'],
        'reg_number': param_dict['reg_number'],
        'board_number': param_dict['board_number'],
        'success_message': success_message,
        'color_mapp': color_mapp
    }
    return render(request, 'docx_create.html', context)

def download_act_docx(request, file_name):
    file_path = os.path.join(settings.MEDIA_ROOT, 'acts', file_name)
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)
    else:
        return HttpResponse('Файл не найден', status=404)

@permission_required('act_creation.add_act')
def application_status_change(request, pk):
    status = Status.objects.get(pk=5)
    application = Application.objects.get(pk=pk)
    if request.method == 'POST':
        application.status = status
        application.save()
    return redirect(reverse('act_creation'))